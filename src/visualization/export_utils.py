"""
Export Utilities for Reports and Data

This module provides functionality to export reports and data in various
formats including CSV, Excel, and PDF.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional
from datetime import datetime
import logging
import os
import json
from pathlib import Path

# Optional imports for Excel and PDF
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import pdfkit
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExportManager:
    """
    Manages export of reports and data to various formats.
    """
    
    def __init__(self, output_dir: str = 'exports'):
        """
        Initialize export manager.
        
        Args:
            output_dir: Base directory for exports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Create subdirectories
        self.csv_dir = self.output_dir / 'csv'
        self.excel_dir = self.output_dir / 'excel'
        self.pdf_dir = self.output_dir / 'pdf'
        
        for dir_path in [self.csv_dir, self.excel_dir, self.pdf_dir]:
            dir_path.mkdir(exist_ok=True)
    
    def export_trades_csv(
        self,
        trades: List[Dict],
        filename: str = 'all_trades.csv'
    ) -> str:
        """
        Export trades to CSV file.
        
        Args:
            trades: List of trade dictionaries
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        # Convert to DataFrame
        trade_df = pd.DataFrame(trades)
        
        # Format columns
        if 'entry_time' in trade_df.columns:
            trade_df['entry_time'] = pd.to_datetime(trade_df['entry_time'])
        if 'exit_time' in trade_df.columns:
            trade_df['exit_time'] = pd.to_datetime(trade_df['exit_time'])
        
        # Reorder columns
        column_order = [
            'trade_id', 'symbol', 'entry_time', 'exit_time', 
            'entry_price', 'exit_price', 'position_size',
            'return', 'pnl', 'confluence_score', 'hold_days',
            'exit_reason', 'max_profit', 'max_loss'
        ]
        
        available_columns = [col for col in column_order if col in trade_df.columns]
        trade_df = trade_df[available_columns]
        
        # Export to CSV
        filepath = self.csv_dir / filename
        trade_df.to_csv(filepath, index=False)
        
        logger.info(f"Exported {len(trades)} trades to {filepath}")
        return str(filepath)
    
    def export_performance_metrics_csv(
        self,
        metrics: Dict[str, Any],
        filename: str = 'performance_metrics.csv'
    ) -> str:
        """
        Export performance metrics to CSV.
        
        Args:
            metrics: Dictionary of metrics
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        # Convert to DataFrame format
        metrics_list = []
        for key, value in metrics.items():
            if isinstance(value, (int, float)):
                metrics_list.append({
                    'Metric': key.replace('_', ' ').title(),
                    'Value': value
                })
        
        metrics_df = pd.DataFrame(metrics_list)
        
        # Export to CSV
        filepath = self.csv_dir / filename
        metrics_df.to_csv(filepath, index=False)
        
        logger.info(f"Exported {len(metrics_list)} metrics to {filepath}")
        return str(filepath)
    
    def export_excel_workbook(
        self,
        trades: List[Dict],
        metrics: Dict[str, Any],
        benchmark_comparison: Optional[Dict[str, Any]] = None,
        filename: str = 'backtest_results.xlsx'
    ) -> Optional[str]:
        """
        Export comprehensive Excel workbook with multiple sheets.
        
        Args:
            trades: List of trades
            metrics: Performance metrics
            benchmark_comparison: Optional benchmark comparison data
            filename: Output filename
            
        Returns:
            Path to exported file or None if Excel not available
        """
        if not EXCEL_AVAILABLE:
            logger.warning("openpyxl not installed. Cannot export Excel file.")
            return None
        
        filepath = self.excel_dir / filename
        
        # Create Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Sheet 1: Trades
            trade_df = pd.DataFrame(trades)
            trade_df.to_excel(writer, sheet_name='Trades', index=False)
            
            # Sheet 2: Performance Metrics
            metrics_df = pd.DataFrame([
                {'Metric': k.replace('_', ' ').title(), 'Value': v}
                for k, v in metrics.items()
                if isinstance(v, (int, float))
            ])
            metrics_df.to_excel(writer, sheet_name='Metrics', index=False)
            
            # Sheet 3: Monthly Summary
            if trades:
                monthly_summary = self._create_monthly_summary(trades)
                monthly_summary.to_excel(writer, sheet_name='Monthly Summary')
            
            # Sheet 4: Benchmark Comparison
            if benchmark_comparison:
                bench_df = pd.DataFrame(benchmark_comparison)
                bench_df.to_excel(writer, sheet_name='Benchmark Comparison', index=False)
        
        # Format the Excel file
        self._format_excel_workbook(filepath)
        
        logger.info(f"Exported Excel workbook to {filepath}")
        return str(filepath)
    
    def _create_monthly_summary(self, trades: List[Dict]) -> pd.DataFrame:
        """Create monthly summary from trades."""
        trade_df = pd.DataFrame(trades)
        trade_df['exit_month'] = pd.to_datetime(trade_df['exit_time']).dt.to_period('M')
        
        monthly = trade_df.groupby('exit_month').agg({
            'return': ['count', 'sum', 'mean'],
            'pnl': 'sum',
            'confluence_score': 'mean'
        })
        
        monthly.columns = ['Trade Count', 'Total Return', 'Avg Return', 'Total PnL', 'Avg Confluence']
        return monthly
    
    def _format_excel_workbook(self, filepath: Path):
        """Apply formatting to Excel workbook."""
        if not EXCEL_AVAILABLE:
            return
        
        wb = openpyxl.load_workbook(filepath)
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for sheet in wb.worksheets:
            # Format headers
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
    
    def export_html_to_pdf(
        self,
        html_content: str,
        filename: str = 'report.pdf',
        options: Optional[Dict[str, Any]] = None
    ) -> Optional[str]:
        """
        Export HTML content to PDF.
        
        Args:
            html_content: HTML string to convert
            filename: Output filename
            options: wkhtmltopdf options
            
        Returns:
            Path to exported file or None if PDF conversion not available
        """
        if not PDF_AVAILABLE:
            logger.warning("pdfkit not installed. Cannot export PDF.")
            return None
        
        filepath = self.pdf_dir / filename
        
        # Default options
        if options is None:
            options = {
                'page-size': 'A4',
                'margin-top': '0.75in',
                'margin-right': '0.75in',
                'margin-bottom': '0.75in',
                'margin-left': '0.75in',
                'encoding': 'UTF-8',
                'no-outline': None
            }
        
        try:
            pdfkit.from_string(html_content, str(filepath), options=options)
            logger.info(f"Exported PDF to {filepath}")
            return str(filepath)
        except Exception as e:
            logger.error(f"Failed to export PDF: {e}")
            return None
    
    def export_confluence_scores_timeseries(
        self,
        confluence_history: pd.DataFrame,
        filename: str = 'confluence_scores_timeseries.csv'
    ) -> str:
        """
        Export confluence score time series.
        
        Args:
            confluence_history: DataFrame with confluence scores over time
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        filepath = self.csv_dir / filename
        confluence_history.to_csv(filepath)
        
        logger.info(f"Exported confluence scores to {filepath}")
        return str(filepath)
    
    def export_json_data(
        self,
        data: Dict[str, Any],
        filename: str = 'strategy_data.json'
    ) -> str:
        """
        Export data as JSON.
        
        Args:
            data: Dictionary to export
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        filepath = self.output_dir / filename
        
        # Convert numpy types for JSON serialization
        def convert_types(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                return float(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, pd.Timestamp):
                return obj.isoformat()
            return obj
        
        # Recursively convert types
        cleaned_data = json.loads(
            json.dumps(data, default=convert_types)
        )
        
        with open(filepath, 'w') as f:
            json.dump(cleaned_data, f, indent=2)
        
        logger.info(f"Exported JSON data to {filepath}")
        return str(filepath)
    
    def create_export_summary(self) -> Dict[str, List[str]]:
        """
        Create summary of all exported files.
        
        Returns:
            Dictionary mapping export type to list of files
        """
        summary = {
            'csv': list(self.csv_dir.glob('*.csv')),
            'excel': list(self.excel_dir.glob('*.xlsx')),
            'pdf': list(self.pdf_dir.glob('*.pdf')),
            'json': list(self.output_dir.glob('*.json'))
        }
        
        # Convert paths to strings
        for key in summary:
            summary[key] = [str(p) for p in summary[key]]
        
        return summary
    
    def export_all(
        self,
        trades: List[Dict],
        metrics: Dict[str, Any],
        confluence_history: Optional[pd.DataFrame] = None,
        benchmark_comparison: Optional[Dict[str, Any]] = None,
        html_report: Optional[str] = None
    ) -> Dict[str, str]:
        """
        Export all available formats.
        
        Args:
            trades: List of trades
            metrics: Performance metrics
            confluence_history: Optional confluence score history
            benchmark_comparison: Optional benchmark comparison
            html_report: Optional HTML report content
            
        Returns:
            Dictionary of format to filepath
        """
        exports = {}
        
        # CSV exports
        exports['trades_csv'] = self.export_trades_csv(trades)
        exports['metrics_csv'] = self.export_performance_metrics_csv(metrics)
        
        if confluence_history is not None:
            exports['confluence_csv'] = self.export_confluence_scores_timeseries(
                confluence_history
            )
        
        # Excel export
        excel_path = self.export_excel_workbook(
            trades, metrics, benchmark_comparison
        )
        if excel_path:
            exports['excel'] = excel_path
        
        # JSON export
        all_data = {
            'metrics': metrics,
            'trades': trades,
            'benchmark_comparison': benchmark_comparison,
            'export_timestamp': datetime.now().isoformat()
        }
        exports['json'] = self.export_json_data(all_data)
        
        # PDF export (if HTML provided)
        if html_report:
            pdf_path = self.export_html_to_pdf(html_report)
            if pdf_path:
                exports['pdf'] = pdf_path
        
        logger.info(f"Completed all exports: {len(exports)} files created")
        return exports